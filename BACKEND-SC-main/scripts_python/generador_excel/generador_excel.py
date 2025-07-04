import sys
import json
import base64
import pandas as pd
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def main():
    try:
        sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')
        data = json.load(sys.stdin)
        data_renglones = data.get("data_renglones", [])
        headers = data.get("headers", {})

        df = pd.DataFrame(data_renglones)
        df.rename(columns=headers, inplace=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]

            max_row = worksheet.max_row
            max_col = worksheet.max_column

            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None:
                    continue
                header_str = str(header_cell.value).strip()
                if (header_str.lower().startswith("id") or 
                    header_str in ["FECHA", "NOMBRE USUARIO", "Valorizado"]):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].hidden = True

            fixed_width_columns = [
                "Cliente", "Descripción (pliego)", 
                "Droga + Presentación (KAIROS)", 
                "Mantenimiento", "Observaciones", "DESCRIPCIÓN"
            ]

            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None or worksheet.column_dimensions[col_letter].hidden:
                    continue
                header_str = str(header_cell.value).strip()
                if header_str in fixed_width_columns:
                    worksheet.column_dimensions[col_letter].width = 35
                else:
                    max_length = max((len(str(cell.value)) for cell in worksheet[col_letter] if cell.value), default=0)
                    worksheet.column_dimensions[col_letter].width = max_length + 2

            white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            grey_fill = PatternFill(fill_type="solid", fgColor="D3D3D3")

            for col_idx in range(1, max_col + 1):
                worksheet.cell(row=1, column=col_idx).fill = white_fill

            editable_headers = ["Costo U.", "Mantenimiento", "Observaciones"]
            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None:
                    continue
                header_str = str(header_cell.value).strip()
                for row in range(2, max_row + 1):
                    current_cell = worksheet.cell(row=row, column=col_idx)
                    current_cell.fill = white_fill if header_str in editable_headers else grey_fill

        output.seek(0)
        encoded_excel = base64.b64encode(output.read()).decode('utf-8')
        print(json.dumps({ "fileName": "datos.xlsx", "contentBase64": encoded_excel }))

    except Exception as e:
        print(json.dumps({ "error": str(e) }), file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
