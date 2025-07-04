import pandas as pd
from openpyxl.utils import get_column_letter

def export_to_excel(data_json, headers_json, output_file="datos.xlsx"):
    print("hola que tal")
    try:
        # 1) Crear DataFrame y renombrar columnas
        df = pd.DataFrame(data_json)
        df.rename(columns=headers_json, inplace=True)
        # 2) Exportar el DataFrame a Excel usando openpyxl
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # 3) Ocultar columnas:
            #    - Si el encabezado contiene "id" (sin distinguir mayúsculas/minúsculas)
            #    - O si es exactamente "FECHA", "NOMBRE USUARIO" o "Valorizado"
            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None: 
                    continue
                header_str = str(header_cell.value).strip()
                if ("id" in header_str.lower() or 
                    header_str in ["FECHA", "NOMBRE USUARIO", "Valorizado"]):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].hidden = True
            
            # 4) Definir lista de columnas con ancho fijo (35)
            fixed_width_columns = [
                "Cliente", "Descripción (pliego)", 
                "Droga + Presentación (KAIROS)", 
                "Mantenimiento", "Observaciones"
            ]
            
            # 5) Ajustar ancho de columnas:
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None:
                    continue
                header_str = str(header_cell.value).strip()
                
                # Si la columna está oculta, no se ajusta ancho
                if worksheet.column_dimensions[col_letter].hidden:
                    continue
                
                if header_str in fixed_width_columns:
                    worksheet.column_dimensions[col_letter].width = 35
                else:
                    # Ajuste automático con padding de 2
                    max_length = 0
                    for cell in worksheet[col_letter]:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    worksheet.column_dimensions[col_letter].width = max_length + 2
            
            # 6) Aplicar colores:
            # Usamos dos fills, blanco y gris.
            from openpyxl.styles import PatternFill
            white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            grey_fill  = PatternFill(fill_type="solid", fgColor="D3D3D3")
            
            # Colorear la fila de encabezados en blanco para todas las columnas.
            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                header_cell.fill = white_fill
            
            # Para los datos (filas 2 a max_row):
            # Las columnas "Costo U.", "Mantenimiento" y "Observaciones" se dejan en blanco;
            # el resto se colorea de gris.
            editable_headers = ["Costo U.", "Mantenimiento", "Observaciones"]
            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None:
                    continue
                header_str = str(header_cell.value).strip()
                col_letter = get_column_letter(col_idx)
                for row in range(2, max_row + 1):
                    current_cell = worksheet.cell(row=row, column=col_idx)
                    if header_str in editable_headers:
                        current_cell.fill = white_fill
                    else:
                        current_cell.fill = grey_fill
        print(f"Datos exportados exitosamente a '{output_file}'")
    except Exception as e:
        print(f"Error al exportar datos: {e}")


# import pandas as pd

# def export_to_excel(data_json, headers_json, output_file="datos.xlsx"):
#     """
#     Exporta datos a un archivo Excel.

#     :param data_json: Lista de diccionarios con los datos de la tabla.
#     :param headers_json: Diccionario con los encabezados de la tabla.
#     :param output_file: Nombre del archivo Excel de salida (por defecto: 'datos.xlsx').
#     """
#     try:
#         # Convertir datos a DataFrame
#         df = pd.DataFrame(data_json)
        
#         # Renombrar columnas según los encabezados proporcionados
#         df.rename(columns=headers_json, inplace=True)
        
#         # Exportar a Excel
#         df.to_excel(output_file, index=False, engine='openpyxl')
#         print(f"Datos exportados exitosamente a {output_file}")
#     except Exception as e:
#         print(f"Error al exportar datos: {e}")


# # Datos de ejemplo
# """ data = [
#     {"renglon": "1", "cantidad": "123", "descripcion": "Prod 1", "codigoTarot": "1747", "productoTarot": "LIDOCAINA 2% S/EPIN AMP X  5ML EV"},
#     {"renglon": "2", "cantidad": "100", "descripcion": "Prod 2", "codigoTarot": "1747", "productoTarot": "LIDOCAINA 2% S/EPIN AMP X  5ML EV"},
#     {"renglon": "3", "cantidad": "200", "descripcion": "Prod 3", "codigoTarot": "1236", "productoTarot": "FLUORESCEINA 10% AMP X 5ML"},
#     {"renglon": "4", "cantidad": "300", "descripcion": "Prod 4", "codigoTarot": "1", "productoTarot": "CLOZAPINA 100MG COMP"},
#     {"renglon": "5", "cantidad": "400", "descripcion": "Prod 5", "codigoTarot": "7741", "productoTarot": "CINTURON P/OSTOMIA 7300 AJUSTABLE"}
# ]

# # Encabezados de ejemplo
# headers = {
#     "renglon": "Renglón",
#     "cantidad": "Cantidad",
#     "descripcion": "Descripción",
#     "codigoTarot": "Código Tarot",
#     "productoTarot": "Producto Tarot"
# } """

# # Exportar a Excel
# #export_to_excel(data, headers, output_file="tabla_productos.xlsx")
