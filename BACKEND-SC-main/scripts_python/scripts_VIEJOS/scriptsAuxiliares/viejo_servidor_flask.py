#from scripts.subir_demanda import analizar_excel
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import io  # Para manejar archivos en memoria como buffer
from scripts.generar_pdf_demanda import generar_PDF
from scriptsAuxiliares.auxiliar_generar_parte import generar_parte_entregas_pdf
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})  

@app.route("/generarpdf", methods=["POST"])
def execute():
    # print("GENERANDO PDF")
    # Extraer los datos de la solicitud
    data = request.json

    data_value = data.get("data_renglones")
    data_cliente = data.get("data_cliente")
    data_entrega = data.get("data_entrega")
    firmas_chequeadas = data.get("firmas_chequeadas")
    total = data.get("total_licitacion")

    if data_cliente != None:
        pdf_buffer = generar_PDF(
            data_value, data_cliente, data_entrega, firmas_chequeadas, total
        )
        # Enviar el PDF como respuesta
        return send_file(
            pdf_buffer,
            as_attachment=False,
            download_name="documento.pdf",
            mimetype="application/pdf",
        )

    return {"columnas_faltantes": "asd"}

@app.route('/generarexcel', methods=['POST'])
def generate_excel():
    try:
        data = request.json
        data_renglones = data.get("data_renglones", [])
        headers = data.get("headers", {})

        df = pd.DataFrame(data_renglones)
        df.rename(columns=headers, inplace=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]

            max_row = worksheet.max_row
            max_col = worksheet.max_column

            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None:
                    continue
                header_str = str(header_cell.value).strip()
                if (header_str.lower().startswith("id") or 
                    header_str in ["FECHA", "NOMBRE USUARIO", "Valorizado"]):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].hidden = True

            fixed_width_columns = [
                "Cliente", "Descripción (pliego)", 
                "Droga + Presentación (KAIROS)", 
                "Mantenimiento", "Observaciones", "DESCRIPCIÓN"
            ]

            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None or worksheet.column_dimensions[col_letter].hidden:
                    continue
                header_str = str(header_cell.value).strip()
                if header_str in fixed_width_columns:
                    worksheet.column_dimensions[col_letter].width = 35
                else:
                    max_length = max((len(str(cell.value)) for cell in worksheet[col_letter] if cell.value), default=0)
                    worksheet.column_dimensions[col_letter].width = max_length + 2

            white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            grey_fill = PatternFill(fill_type="solid", fgColor="D3D3D3")

            for col_idx in range(1, max_col + 1):
                worksheet.cell(row=1, column=col_idx).fill = white_fill

            editable_headers = ["Costo U.", "Mantenimiento", "Observaciones"]
            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                if header_cell.value is None:
                    continue
                header_str = str(header_cell.value).strip()
                for row in range(2, max_row + 1):
                    current_cell = worksheet.cell(row=row, column=col_idx)
                    current_cell.fill = white_fill if header_str in editable_headers else grey_fill

        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="datos.xlsx"
        )
    except Exception as e:
        print(f"Error al exportar datos: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/generarparteentregas", methods=["POST"])
def generar_parte_entregas():
    try:
        data = request.json
        json_data = data.get("json_data")
        entregas = data.get("entregas")

        if not json_data or not entregas:
            return {"error": "Faltan json_data o entregas"}, 400

        pdf_buffer = generar_parte_entregas_pdf(json_data, entregas)

        return send_file(
            pdf_buffer,
            as_attachment=False,
            download_name="parte_entregas.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        print(f"Error al generar parte entregas: {e}")
        return {"error": str(e)}, 500


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)

